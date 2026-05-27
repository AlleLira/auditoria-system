from openpyxl import Workbook

from openpyxl.styles import (

    Font,
    PatternFill,
    Alignment,
    Border,
    Side

)

from openpyxl.utils import get_column_letter

import pandas as pd


def gerar_relatorio_excel(
    df,
    data_inicial
):

    if df.empty:

        return None


    wb = Workbook()

    ws = wb.active


    mes = data_inicial.strftime("%m")
    ano = data_inicial.strftime("%Y")


    meses = {

        "01":"JANEIRO",
        "02":"FEVEREIRO",
        "03":"MARÇO",
        "04":"ABRIL",
        "05":"MAIO",
        "06":"JUNHO",
        "07":"JULHO",
        "08":"AGOSTO",
        "09":"SETEMBRO",
        "10":"OUTUBRO",
        "11":"NOVEMBRO",
        "12":"DEZEMBRO"

    }


    titulo = (

        f"PENDÊNCIAS AUDITORIA "

        f"{meses[mes]}/{ano}"

    )


    ws.merge_cells("A1:E1")

    ws["A1"] = titulo

    ws["A1"].font = Font(
        bold=True,
        size=16
    )

    ws["A1"].alignment = Alignment(
        horizontal="center"
    )


    headers = [

        "PDV",
        "Consultor",
        "Valor",
        "20%",
        "Valor a ser debitado"

    ]


    linha_header = 3


    for col, header in enumerate(headers, start=1):

        cell = ws.cell(
            row=linha_header,
            column=col
        )

        cell.value = header

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(

            start_color="F4B400",

            end_color="F4B400",

            fill_type="solid"

        )

        cell.alignment = Alignment(
            horizontal="center"
        )


    agrupado = (

        df.groupby(

            ["loja","consultor"]

        )["valor"]

        .sum()

        .reset_index()

    )


    linha = 4


    total_valor = 0
    total_20 = 0
    total_debito = 0


    borda = Border(

        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')

    )


    for _, row in agrupado.iterrows():

        valor = float(
            row["valor"]
        )

        vinte = valor * 0.20

        debito = min(
            vinte,
            50
        )


        total_valor += valor
        total_20 += vinte
        total_debito += debito


        dados = [

            row["loja"],
            row["consultor"],
            valor,
            vinte,
            debito

        ]


        for col, valor_celula in enumerate(dados, start=1):

            cell = ws.cell(
                row=linha,
                column=col
            )

            cell.value = valor_celula

            cell.border = borda

            if col >= 3:

                cell.number_format = 'R$ #,##0.00'


        linha += 1


    ws.merge_cells(

        start_row=linha,
        start_column=1,
        end_row=linha,
        end_column=2

    )


    total_cell = ws.cell(
        row=linha,
        column=1
    )

    total_cell.value = "TOTAL"

    total_cell.font = Font(
        bold=True,
        color="FFFFFF"
    )

    total_cell.fill = PatternFill(

        start_color="F4B400",

        end_color="F4B400",

        fill_type="solid"

    )

    total_cell.alignment = Alignment(
        horizontal="center"
    )


    totais = [

        total_valor,
        total_20,
        total_debito

    ]


    for i, valor_total in enumerate(totais, start=3):

        cell = ws.cell(
            row=linha,
            column=i
        )

        cell.value = valor_total

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(

            start_color="F4B400",

            end_color="F4B400",

            fill_type="solid"

        )

        cell.number_format = 'R$ #,##0.00'


    larguras = {

        1:15,
        2:35,
        3:15,
        4:15,
        5:25

    }


    for col, largura in larguras.items():

        ws.column_dimensions[
            get_column_letter(col)
        ].width = largura


    caminho = "relatorio_auditoria.xlsx"

    wb.save(caminho)

    return caminho